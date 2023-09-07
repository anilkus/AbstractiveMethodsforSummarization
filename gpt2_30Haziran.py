from transformers import GPT2Tokenizer, GPT2LMHeadModel
from nltk.translate.bleu_score import corpus_bleu, sentence_bleu


def abstractive_summarization(text):
    # GPT-2 tokenizer'ı yükle
    tokenizer = GPT2Tokenizer.from_pretrained('gpt2')

    # Metni tokenize et
    inputs = tokenizer.encode(text, return_tensors='pt', max_length=512, truncation=True)

    # GPT-2 modelini yükle
    model = GPT2LMHeadModel.from_pretrained('gpt2')

    # Özetleme işlemini gerçekleştir
    outputs = model.generate(inputs, max_length=1024, min_length=30, num_beams=4, early_stopping=True)

    # Özet metni decode et
    summary = tokenizer.decode(outputs[0], skip_special_tokens=True)

    return summary


# Dosyanın içeriğini oku
with open("C:/Users/LENOVO/Desktop/TEZ/Veri Ön İşleme/Abstracts/z_Uludağ Üniversitesi Tıp Fakültesi Dergisi/979546-1911832.txt", 'r',encoding="utf-8") as text2:
    abstract=text2.read()
with open("C:/Users/LENOVO/Desktop/TEZ/Veri Ön İşleme/İşlenecek Kısım/z_Uludağ Üniversitesi Tıp Fakültesi Dergisi/979546-1911832.txt", 'r',encoding="utf-8") as text:
    islenecek=text.read()
    text=islenecek

# Metnin özetini al
summary = abstractive_summarization(text)

# Özeti ekrana yazdır
print("Metin Özeti:")
print(summary)
from rouge import Rouge
ROUGE = Rouge()

print(ROUGE.get_scores(summary, abstract))

gercekmetinuzunlugu=len(text)
özetuzunlugu=len(summary)
orjinalozet=len(abstract)


'''

from openpyxl import load_workbook
wb = load_workbook('C:/Users/LENOVO/Desktop/TEZ/Veri Ön İşleme/Sonuçlar/GPT.xlsx')
ws = wb['Sayfa1']
yayın1 = "Uludağ "
yayın2 = "1"

recall = 0.12
precision = 0.19

ws.cell(row=4,column=3).value = yayın1
ws.cell(row=4,column=4).value = yayın2
ws.cell(row=4,column=5).value = gercekmetinuzunlugu
ws.cell(row=4,column=6).value = özetuzunlugu
ws.cell(row=4,column=8).value = orjinalozet
ws.cell(row=4,column=9).value = recall
ws.cell(row=4,column=10).value = precision

wb.save('C:/Users/LENOVO/Desktop/TEZ/Veri Ön İşleme/Sonuçlar/GPT.xlsx')


ozet=summary
with open("C:/Users/LENOVO/Desktop/TEZ/Veri Ön İşleme/Çalışmanın Özetleri/GPT/z_Uludağ Üniversitesi Tıp Fakültesi Dergisi/971366-1880431.txt", "w",encoding="utf-8") as file:
    file.write(ozet)

'''
